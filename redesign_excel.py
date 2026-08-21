from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

excel_file = r"Modelos_Recria_Engorde_2026.xlsx"
wb = load_workbook(excel_file)
ws = wb["Parametros a actualizar"]

# Encontrar última fila
last_row = 36

row = last_row + 2

# Título
ws[f'B{row}'] = "6. SITUACION ACTUAL — ¿Dónde estamos?"
ws[f'B{row}'].font = Font(bold=True, size=12)
row += 2

# Cuál camino
ws[f'B{row}'] = "¿Cuál camino eligió al inicio?"
ws[f'C{row}'] = 1
ws[f'C{row}'].font = Font(color="0000ff")
ws[f'D{row}'] = "(1=Venta oct, 2=Cap, 3=Corral, 4=Pastura, 5=Recría+corral)"
ws[f'D{row}'].font = Font(italic=True, size=9)
row += 1

# Fecha hoy
ws[f'B{row}'] = "Fecha de hoy"
ws[f'C{row}'] = "=TODAY()"
ws[f'C{row}'].font = Font(color="0000ff")
row += 1

# Margen acumulado
ws[f'B{row}'] = "Margen acumulado hasta hoy"
ws[f'C{row}'] = 0
ws[f'C{row}'].font = Font(color="0000ff")

wb.save(excel_file)
print("✅ Excel actualizado con sección 'Situación Actual'")
